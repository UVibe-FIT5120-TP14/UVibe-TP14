"""
cancer_incident_wrangler.py
----------------------------
Unified wrangler for AIHW cancer incidence Excel files.

Handles two datasets:

  State mode  (--mode state)
    Source: aihw-can-122-CDiA-2023-Book-7-Cancer-incidence-by-state-and-territory.xlsx
    Sheet:  Table S7.1
    Columns: Data type | Cancer group/site | Year | Sex | State or Territory | Count
    Table:  state_cancer_incidents

  Age mode  (--mode age)
    Source: aihw-can-122-CDiA-2023-Book-1a-Cancer-incidence-age-standardised-rates-5-year-age-groups.xlsx
    Sheet:  Table S1a.1
    Columns: Data type | Cancer group/site | Year | Sex | Age group (years) | Count
    Table:  age_cancer_incidents

By default (no --mode flag) both tables are loaded sequentially.

Suppressed / unavailable values ('n.p.', '. .', '') are stored as NULL.

Usage:
    cd backend
    python scripts/cancer_incident_wrangler.py            # run both
    python scripts/cancer_incident_wrangler.py --mode state
    python scripts/cancer_incident_wrangler.py --mode age
    python scripts/cancer_incident_wrangler.py --data-dir ../data/cancer_incidents
"""

import argparse
from dataclasses import dataclass
import os
import sys
from typing import Callable

import openpyxl
from sqlalchemy.orm import Session

from database import SessionLocal, engine
from models import Base, StateCancerIncident, AgeCancerIncidents 

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))


SKIN_CANCER_TYPES = {
    "Melanoma of the skin",
    "Non-melanoma skin cancer (rare types)",
}

# Values used in the spreadsheet to indicate suppressed or missing data
SUPPRESSED_VALUES = {"n.p.", ". .", "..", ""}

def _parse_int(value) -> int | None:
    """Return an integer, or None for suppressed / missing values."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if text in SUPPRESSED_VALUES:
        return None
    try:
        return int(float(text))
    except (ValueError, TypeError):
        return None


@dataclass
class WranglerConfig:
    """Everything that differs between the two source files."""
    label: str                              # human-readable name for logging
    filename: str                           # Excel filename
    sheet_name: str                         # sheet tab to read
    data_start_row: int                     # 1-indexed first data row (rows before are header/metadata)
    valid_data_type: str                    # value in column 0 to keep (e.g. "Incidence" or "Actual")
    dimension_col: int                      # column index (0-based) holding state or age group
    exclude_dimension: set[str]             # dimension values to skip (e.g. aggregate rows)
    model_factory: Callable                 # callable(cancer_type, year, sex, dimension, count) -> ORM obj
    model_class: type                       # ORM class (used to clear table before insert)


def _wrangle(cfg: WranglerConfig, data_dir: str, db: Session) -> int:
    filepath = os.path.join(data_dir, cfg.filename)
    if not os.path.exists(filepath):
        print(f"[ERROR] File not found: {filepath}")
        sys.exit(1)

    print(f"\n--- {cfg.label} ---")
    print(f"Opening: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if cfg.sheet_name not in wb.sheetnames:
        print(f"[ERROR] Sheet '{cfg.sheet_name}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[cfg.sheet_name]

    # Clear existing rows so re-runs are idempotent
    deleted = db.query(cfg.model_class).delete()
    db.commit()
    print(f"Cleared {deleted} existing rows from '{cfg.model_class.__tablename__}'.")

    records = []
    skipped = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Skip header / metadata rows at the top
        if row_idx < cfg.data_start_row:
            continue

        # Minimum columns: data_type(0), cancer_type(1), year(2), sex(3), dimension(4), count(5)
        if len(row) < 6:
            skipped += 1
            continue

        data_type   = str(row[0]).strip() if row[0] is not None else ""
        cancer_type = str(row[1]).strip() if row[1] is not None else ""
        year_val    = row[2]
        sex         = str(row[3]).strip() if row[3] is not None else ""
        dimension   = str(row[cfg.dimension_col]).strip() if row[cfg.dimension_col] is not None else ""
        count_val   = row[5]

        # Only keep rows with the expected data type tag
        if data_type != cfg.valid_data_type:
            skipped += 1
            continue

        # Skip aggregate / unwanted dimension values (e.g. "Australia", "All ages")
        if dimension.lower() in cfg.exclude_dimension:
            skipped += 1
            continue

        # Require core fields
        if not cancer_type or not sex or not dimension:
            skipped += 1
            continue

        # Only skin-cancer-related rows
        if cancer_type not in SKIN_CANCER_TYPES:
            skipped += 1
            continue

        # Only Males / Females (skip combined "Persons" totals)
        if sex == "Persons":
            skipped += 1
            continue

        # Skip rows with suppressed counts
        count_str = str(count_val).strip() if count_val is not None else ""
        if count_str in SUPPRESSED_VALUES or count_val is None:
            skipped += 1
            continue

        year = _parse_int(year_val)
        if year is None:
            skipped += 1
            continue

        count = _parse_int(count_val)

        records.append(cfg.model_factory(
            cancer_type=cancer_type,
            year=year,
            sex=sex,
            dimension=dimension,
            count=count,
        ))

    print(f"Inserting {len(records)} records ({skipped} rows skipped)...")
    db.bulk_save_objects(records)
    db.commit()
    print(f"Done — {cfg.label}.")
    return len(records)


def _state_config() -> WranglerConfig:
    return WranglerConfig(
        label="State/Territory cancer incidents (Book 7)",
        filename="aihw-can-122-CDiA-2023-Book-7-Cancer-incidence-by-state-and-territory.xlsx",
        sheet_name="Table S7.1",
        data_start_row=6,
        valid_data_type="Incidence",
        dimension_col=4,
        exclude_dimension={"australia"},  # lowercase comparison
        model_class=StateCancerIncident,
        model_factory=lambda cancer_type, year, sex, dimension, count: StateCancerIncident(
            cancer_type=cancer_type,
            year=year,
            sex=sex,
            state=dimension,
            count=count,
        ),
    )


def _age_config() -> WranglerConfig:
    return WranglerConfig(
        label="Age-group cancer incidents (Book 1a)",
        filename="aihw-can-122-CDiA-2023-Book-1a-Cancer-incidence-age-standardised-rates-5-year-age-groups.xlsx",
        sheet_name="Table S1a.1",
        data_start_row=6,
        valid_data_type="Actual",
        dimension_col=4,
        exclude_dimension={"all ages combined"},
        model_class=AgeCancerIncidents,
        model_factory=lambda cancer_type, year, sex, dimension, count: AgeCancerIncidents(
            cancer_type=cancer_type,
            year=year,
            sex=sex,
            age_group=dimension,
            count=count,
        ),
    )


def wrangle_cancer_incident_history(base_dir: str) -> None:
    # Ensure all tables exist before we write
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    data_dir = os.path.join(base_dir, "cancer_incidents")

    try:
        configs = [_state_config(), _age_config()]
        total = 0
        for cfg in configs:
            total += _wrangle(cfg, data_dir=data_dir, db=db)

        print(f"\nTotal records inserted: {total}")
    finally:
        db.close()
